import json
import os
from openpyxl import load_workbook

class ReadExcel(object):
    def __init__(self,filepath,sheet_name):
        self.wb = load_workbook(filepath)
        self.sh = self.wb[sheet_name]

    def read_titles(self):
        titles = []
        for item in list(self.sh.rows)[0]:
            titles.append(item.value)
        return titles

    def read_data(self):
        all_data = []
        titles = self.read_titles()
        for item in list(self.sh.rows)[1:]:
            values = []
            for val in item:
                values.append(val.value)
            res = dict(zip(titles,values))
            res["request_data"] = json.loads(res["request_data"])
            # eval()如果有None会报错
            res["expected"] = eval(res["expected"])
            all_data.append(res)
        return all_data

    def close_file(self):
        self.wb.close()

if __name__ == "__main__":
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)),'login_data.xlsx')
    exc = ReadExcel(filepath,"login_data")
    case = exc.read_data()
    exc.close_file()
    for cas in case:
        print(cas)




# # 获取文件路径
# file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),'login_data.xlsx')
# # print(file_path)
# # 生成workbook对象
# wb = load_workbook(file_path)
# # 根据表单名称取得表单
# sh = wb['login_data']
# # print(sh)
# # 获取单元格对象
# cel = sh.cell(2,2)
# print(cel.value)
# print(sh.max_row)  # 总行数
# print(sh.max_column)  # 总列数

# 修改值
# sh.cell(4,1).value = "admin22"
# print(sh.cell(4,1).value)
# 另存为
# wb.save("save_another_excel.xlsx")

# 按行读取
# print(list(sh.rows))
# for item in list(sh.rows):
#     # print(item)
#     for ce in item:
#         print(ce.value)
# titles = []
# for item in list(sh.rows)[0]:
#     titles.append(item.value)
# print(titles)
# data_lists = []
# for item in list(sh.rows)[1:]:
#     value_dict = {}
#     for index in range(len(item)):
#         print(index,item[index],item[index].value)
#         value_dict[titles[index]] = item[index].value
#     print(value_dict)
#     data_lists.append(value_dict)
# print(data_lists)
# titles = []
# for item in list(sh.rows)[0]:
#     titles.append(item.value)
# all_data = []
# for item in list(sh.rows)[1:]:
#     values = []
#     for val in item:  # 遍历数据行
#         values.append(val.value)
#     res = dict(zip(titles,values))  # titles和数据打包成字典
#     res['check'] = eval(res['check'])  # 将check的字符串转换为字典对象
#     all_data.append(res)
# print(all_data)
# wb.close()
